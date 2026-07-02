from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import AlphaExpression, SimulationSettings


class ExcelInputError(ValueError):
    pass


SETTING_COLUMNS = set(SimulationSettings.__dataclass_fields__.keys())
INT_SETTING_COLUMNS = {"delay", "decay"}
FLOAT_SETTING_COLUMNS = {"truncation"}
BOOL_SETTING_COLUMNS = {"visualization"}


def _coerce_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.lower() == "true":
            return True
        if stripped.lower() == "false":
            return False
        return stripped
    return value


def _coerce_setting_value(value: Any, *, header: str, row_number: int, sheet_name: str) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        value = value.strip()
    if value == "":
        return ""

    try:
        if header in INT_SETTING_COLUMNS:
            if isinstance(value, bool):
                raise ValueError
            if isinstance(value, float) and not value.is_integer():
                raise ValueError
            return int(value)
        if header in FLOAT_SETTING_COLUMNS:
            if isinstance(value, bool):
                raise ValueError
            return float(value)
        if header in BOOL_SETTING_COLUMNS:
            if isinstance(value, bool):
                return value
            if isinstance(value, str) and value.lower() in {"true", "false"}:
                return value.lower() == "true"
            raise ValueError
    except (TypeError, ValueError) as exc:
        raise ExcelInputError(
            f"Invalid value for sheet '{sheet_name}', row {row_number}, column '{header}'."
        ) from exc

    return value


def read_excel_expressions(path: str | Path, *, sheet_name: str | None = None) -> list[AlphaExpression]:
    workbook = load_workbook(Path(path), data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            available_sheets = ", ".join(workbook.sheetnames)
            raise ExcelInputError(
                f"Excel sheet '{sheet_name}' was not found. Available sheets: {available_sheets}."
            )
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    active_sheet_name = str(sheet.title)
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ExcelInputError("Excel file is empty.")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    header_to_index = {header: index for index, header in enumerate(headers) if header}
    if "expression" not in header_to_index:
        raise ExcelInputError("Excel input must contain an expression column.")

    output: list[AlphaExpression] = []
    for excel_row_number, row in enumerate(rows[1:], start=2):
        expression = _coerce_value(row[header_to_index["expression"]])
        if not expression:
            continue
        row_id = ""
        if "id" in header_to_index:
            row_id = str(_coerce_value(row[header_to_index["id"]]))
        if not row_id:
            row_id = f"row-{excel_row_number}"

        overrides: dict[str, Any] = {}
        metadata: dict[str, Any] = {"excel_row": excel_row_number}
        for header, index in header_to_index.items():
            raw_value = row[index] if index < len(row) else ""
            value = _coerce_value(raw_value)
            if header in SETTING_COLUMNS:
                overrides[header] = _coerce_setting_value(
                    raw_value,
                    header=header,
                    row_number=excel_row_number,
                    sheet_name=active_sheet_name,
                )
            elif header not in {"id", "expression"}:
                metadata[header] = value

        output.append(
            AlphaExpression(
                row_id=row_id,
                expression=str(expression),
                settings=SimulationSettings.from_overrides(overrides),
                metadata=metadata,
            )
        )
    return output
