from __future__ import annotations

import csv
import json
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
EXAMPLES = ROOT / "examples"
DATA_DIR = EXAMPLES / "data"
EXPECTED_DIR = EXAMPLES / "expected"
RUNS_DIR = EXAMPLES / "runs"
FIXED_WORKBOOK_DATETIME = datetime(2026, 1, 1, 0, 0, 0)
FIXED_WORKBOOK_TIMESTAMP = "2026-01-01T00:00:00Z"
FIXED_ZIP_DATETIME = (2026, 1, 1, 0, 0, 0)
CORE_XML_NAMESPACES = {
    "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
    "dc": "http://purl.org/dc/elements/1.1/",
    "dcterms": "http://purl.org/dc/terms/",
    "xsi": "http://www.w3.org/2001/XMLSchema-instance",
}

for prefix, uri in CORE_XML_NAMESPACES.items():
    ET.register_namespace(prefix, uri)

SUMMARY_FIELDS = [
    "row_id",
    "alpha_hash",
    "status",
    "alpha_id",
    "simulation_location",
    "sharpe",
    "fitness",
    "returns",
    "turnover",
    "drawdown",
    "margin",
    "longCount",
    "shortCount",
    "failed_checks",
    "warning_checks",
    "pending_checks",
    "error",
]

STANDARD_HEADERS = [
    "id",
    "expression",
    "region",
    "universe",
    "delay",
    "decay",
    "neutralization",
    "truncation",
    "nanHandling",
    "language",
    "visualization",
    "theme",
]

ALPHA_ROWS = [
    ("tutorial-001", "rank(ts_mean(volume, 10))", "liquidity"),
    ("tutorial-002", "rank(ts_mean(volume, 20))", "liquidity"),
    ("tutorial-003", "rank(ts_mean(returns, 5))", "momentum"),
    ("tutorial-004", "rank(ts_std_dev(returns, 20))", "risk"),
]

LIVE_ROWS = [
    ("live-001", "rank(ts_mean(volume, 10))", "liquidity"),
    ("live-002", "rank(ts_mean(returns, 5))", "momentum"),
    ("live-003", "rank(ts_std_dev(returns, 20))", "risk"),
    ("live-004", "rank(ts_delta(close, 5))", "price"),
]

DUPLICATE_ROWS = [
    ("duplicate-001", "rank(ts_mean(volume, 10))", "liquidity"),
    ("duplicate-002", "rank(ts_mean(returns, 5))", "momentum"),
]

FALLBACK_ROWS = [
    (f"fallback-{index:03d}", f"rank(ts_mean(volume, {index + 4}))", "fallback")
    for index in range(1, 9)
]

RECORDSET_ROWS = [
    ("recordset-001", "rank(ts_mean(volume, 10))", "liquidity"),
    ("recordset-002", "rank(ts_mean(returns, 5))", "momentum"),
]

TUTORIALS = [
    "Tutorial 0 - Start Here For Beginners.ipynb",
    "Tutorial 1 - Installation And Project Tour.ipynb",
    "Tutorial 2 - Login And Persona Verification.ipynb",
    "Tutorial 3 - Excel Alpha Queue And Payloads.ipynb",
    "Tutorial 4 - Live Excel Batch Simulation.ipynb",
    "Tutorial 5 - Batch Fallback Timeouts And Retry Queue.ipynb",
    "Tutorial 6 - Duplicate Cache And Re-Runs.ipynb",
    "Tutorial 7 - Results Raw Logs And Recordsets.ipynb",
    "Tutorial 8 - Python API Workflow.ipynb",
]

OBSOLETE_NOTEBOOKS = [
    "Tutorial 1 - Excel Batch Alpha Simulation.ipynb",
]


def _normalize_core_properties(data: bytes) -> bytes:
    root = ET.fromstring(data)
    dcterms = CORE_XML_NAMESPACES["dcterms"]
    for tag_name in ("created", "modified"):
        element = root.find(f"{{{dcterms}}}{tag_name}")
        if element is not None:
            element.text = FIXED_WORKBOOK_TIMESTAMP
    return ET.tostring(root, encoding="utf-8", xml_declaration=False)


def _canonicalize_xml(data: bytes) -> bytes:
    return ET.canonicalize(data.decode("utf-8")).encode("utf-8")


def _normalize_xlsx_zip(path: Path) -> None:
    with tempfile.NamedTemporaryFile(dir=path.parent, delete=False) as temp_file:
        temp_path = Path(temp_file.name)

    try:
        with zipfile.ZipFile(path, "r") as source:
            with zipfile.ZipFile(
                temp_path,
                "w",
                compression=zipfile.ZIP_DEFLATED,
                compresslevel=9,
            ) as target:
                for source_info in sorted(source.infolist(), key=lambda info: info.filename):
                    data = source.read(source_info.filename)
                    if source_info.filename == "docProps/core.xml":
                        data = _normalize_core_properties(data)
                    if source_info.filename.endswith((".xml", ".rels")):
                        data = _canonicalize_xml(data)
                    target_info = zipfile.ZipInfo(source_info.filename, FIXED_ZIP_DATETIME)
                    target_info.compress_type = zipfile.ZIP_DEFLATED
                    target_info.create_system = 0
                    target_info.external_attr = 0o600 << 16
                    target.writestr(target_info, data)
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _write_workbook(path: Path, rows: Iterable[tuple[str, str, str]], *, invalid: bool = False) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.properties.created = FIXED_WORKBOOK_DATETIME
    workbook.properties.modified = FIXED_WORKBOOK_DATETIME
    sheet = workbook.active
    sheet.title = "alphas"
    headers = ["id", "region", "universe"] if invalid else STANDARD_HEADERS
    sheet.append(headers)
    for row_id, expression, theme in rows:
        if invalid:
            sheet.append([row_id, "USA", "TOP3000"])
            continue
        sheet.append(
            [
                row_id,
                expression,
                "USA",
                "TOP3000",
                1,
                5,
                "SUBINDUSTRY",
                0.08,
                "OFF",
                "FASTEXPR",
                False,
                theme,
            ]
        )
    workbook.save(path)
    _normalize_xlsx_zip(path)


def build_workbooks() -> None:
    _write_workbook(DATA_DIR / "tutorial_03_mixed_settings.xlsx", ALPHA_ROWS)
    _write_workbook(DATA_DIR / "tutorial_03_invalid_missing_expression.xlsx", ALPHA_ROWS[:1], invalid=True)
    _write_workbook(DATA_DIR / "tutorial_04_live_alphas.xlsx", LIVE_ROWS)
    _write_workbook(DATA_DIR / "tutorial_05_fallback_alphas.xlsx", FALLBACK_ROWS)
    _write_workbook(DATA_DIR / "tutorial_06_duplicate_alphas.xlsx", DUPLICATE_ROWS)
    _write_workbook(DATA_DIR / "tutorial_07_recordset_alphas.xlsx", RECORDSET_ROWS)
    _write_workbook(DATA_DIR / "tutorial_08_api_alphas.xlsx", ALPHA_ROWS[:3])
    legacy_path = DATA_DIR / "tutorial_01_alphas.xlsx"
    if legacy_path.exists():
        legacy_path.unlink()


def _summary_row(
    *,
    row_id: str,
    index: int,
    alpha_hash: str,
    status: str = "complete",
    alpha_id: str | None = None,
    simulation_location: str | None = None,
    error: str = "",
) -> dict[str, str]:
    alpha_id = alpha_id if alpha_id is not None else f"tutorial-alpha-{index}"
    simulation_location = simulation_location if simulation_location is not None else f"/simulations/tutorial-{index}"
    return {
        "row_id": row_id,
        "alpha_hash": alpha_hash,
        "status": status,
        "alpha_id": alpha_id,
        "simulation_location": simulation_location,
        "sharpe": f"{1.0 + index / 10:.2f}" if status == "complete" else "",
        "fitness": f"{0.5 + index / 10:.2f}" if status == "complete" else "",
        "returns": f"{0.03 + index / 100:.2f}" if status == "complete" else "",
        "turnover": f"{0.10 + index / 100:.2f}" if status == "complete" else "",
        "drawdown": f"{0.02 + index / 100:.2f}" if status == "complete" else "",
        "margin": "",
        "longCount": "",
        "shortCount": "",
        "failed_checks": "",
        "warning_checks": "",
        "pending_checks": "",
        "error": error,
    }


def _write_summary(path: Path, rows: list[dict[str, str]]) -> None:
    EXPECTED_DIR.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=SUMMARY_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def build_expected_outputs() -> None:
    _write_summary(
        EXPECTED_DIR / "tutorial_04_live_offline_summary.csv",
        [
            _summary_row(
                row_id=row_id,
                index=index,
                alpha_hash=f"tutorial-hash-{index}",
                simulation_location="/simulations/tutorial-1",
            )
            for index, (row_id, _expression, _theme) in enumerate(LIVE_ROWS, start=1)
        ],
    )
    _write_summary(
        EXPECTED_DIR / "tutorial_05_fallback_summary.csv",
        [
            *[
                _summary_row(
                    row_id=row_id,
                    index=index,
                    alpha_hash=f"fallback-hash-{index}",
                    simulation_location="/simulations/fallback-1",
                )
                for index, (row_id, _expression, _theme) in enumerate(FALLBACK_ROWS[:4], start=1)
            ],
            *[
                _summary_row(
                    row_id=row_id,
                    index=index,
                    alpha_hash=f"fallback-hash-{index}",
                    status="pending_timeout",
                    alpha_id="",
                    simulation_location="/simulations/fallback-timeout",
                    error="pending_timeout",
                )
                for index, (row_id, _expression, _theme) in enumerate(FALLBACK_ROWS[4:], start=5)
            ],
        ],
    )
    _write_summary(
        EXPECTED_DIR / "tutorial_06_second_run_summary.csv",
        [
            _summary_row(
                row_id=row_id,
                index=index,
                alpha_hash=f"duplicate-hash-{index}",
                status="skipped_duplicate",
                alpha_id=f"duplicate-alpha-{index}",
                simulation_location="",
            )
            for index, (row_id, _expression, _theme) in enumerate(DUPLICATE_ROWS, start=1)
        ],
    )
    _write_summary(
        EXPECTED_DIR / "tutorial_07_recordset_summary.csv",
        [
            _summary_row(row_id=row_id, index=index, alpha_hash=f"recordset-hash-{index}", alpha_id=f"recordset-alpha-{index}")
            for index, (row_id, _expression, _theme) in enumerate(RECORDSET_ROWS, start=1)
        ],
    )
    legacy_summary = EXPECTED_DIR / "tutorial_01_summary.csv"
    if legacy_summary.exists():
        legacy_summary.unlink()


def markdown_cell(source: str) -> dict[str, Any]:
    return {
        "cell_type": "markdown",
        "metadata": {},
        "source": [line + "\n" for line in source.splitlines()],
    }


def code_cell(source: str) -> dict[str, Any]:
    return {
        "cell_type": "code",
        "execution_count": None,
        "metadata": {},
        "outputs": [],
        "source": [line + "\n" for line in source.splitlines()],
    }


def _write_notebook(filename: str, cells: list[dict[str, Any]]) -> None:
    EXAMPLES.mkdir(parents=True, exist_ok=True)
    notebook = {
        "cells": cells,
        "metadata": {
            "kernelspec": {
                "display_name": "Python 3",
                "language": "python",
                "name": "python3",
            },
            "language_info": {
                "name": "python",
                "pygments_lexer": "ipython3",
            },
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    (EXAMPLES / filename).write_text(
        json.dumps(notebook, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


COMMON_SETUP = """from __future__ import annotations

import csv
import json
import os
import shutil
import sqlite3
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

RUN_LIVE = os.getenv("BRAIN_SIM_RUN_LIVE") == "1"
CWD = Path.cwd().resolve()
ROOT = CWD if (CWD / "examples").exists() else CWD.parent if CWD.name == "examples" else CWD
EXAMPLE_DIR = ROOT / "examples"
DATA_DIR = EXAMPLE_DIR / "data"
EXPECTED_DIR = EXAMPLE_DIR / "expected"
RUNS_DIR = EXAMPLE_DIR / "runs"
RUNS_DIR.mkdir(parents=True, exist_ok=True)
print(f"Repository root: {ROOT}")
print(f"Live execution enabled: {RUN_LIVE}")"""

FAKE_CLIENTS = """@dataclass(frozen=True)
class TutorialRateLimitState:
    limit: int | None
    remaining: int | None
    reset_seconds: int | None


@dataclass(frozen=True)
class TutorialSubmitResult:
    status_code: int
    location: str
    body: Any
    headers: dict[str, str]
    rate_limit: TutorialRateLimitState


@dataclass(frozen=True)
class TutorialPollResult:
    status: str
    body: Any
    events: list[dict[str, Any]]


class CompleteFakeBrainClient:
    def __init__(self, alpha_prefix: str = "tutorial-alpha", location_prefix: str = "/simulations/tutorial") -> None:
        self.alpha_prefix = alpha_prefix
        self.location_prefix = location_prefix
        self.submit_count = 0
        self.locations: dict[str, list[str]] = {}

    def submit(self, payload: dict[str, Any] | list[dict[str, Any]]) -> TutorialSubmitResult:
        self.submit_count += 1
        payloads = payload if isinstance(payload, list) else [payload]
        alpha_ids = [f"{self.alpha_prefix}-{self.submit_count + offset}" for offset in range(len(payloads))]
        location = f"{self.location_prefix}-{self.submit_count}"
        self.locations[location] = alpha_ids
        return TutorialSubmitResult(
            status_code=201,
            location=location,
            body={"id": location.rsplit("/", 1)[-1]},
            headers={"Location": location, "X-Ratelimit-Remaining": "999"},
            rate_limit=TutorialRateLimitState(limit=None, remaining=999, reset_seconds=None),
        )

    def poll(self, location: str, *, timeout_seconds: float) -> TutorialPollResult:
        alpha_ids = self.locations[location]
        body: Any = {"alpha": alpha_ids[0]} if len(alpha_ids) == 1 else {"alphas": alpha_ids}
        return TutorialPollResult(status="complete", body=body, events=[{"body": body}])

    def fetch_alpha(self, alpha_id: str) -> dict[str, Any]:
        index = int(alpha_id.rsplit("-", 1)[-1])
        return {
            "id": alpha_id,
            "status": "UNSUBMITTED",
            "is": {
                "sharpe": f"{1.0 + index / 10:.2f}",
                "fitness": f"{0.5 + index / 10:.2f}",
                "returns": f"{0.03 + index / 100:.2f}",
                "turnover": f"{0.10 + index / 100:.2f}",
                "drawdown": f"{0.02 + index / 100:.2f}",
                "checks": [
                    {"name": "LOW_SHARPE", "result": "WARNING"} if index == 2 else {"name": "SELF_CORRELATION", "result": "PASS"}
                ],
            },
        }

    def fetch_recordset(self, alpha_id: str, recordset_name: str) -> dict[str, Any]:
        return {
            "alpha_id": alpha_id,
            "recordset": recordset_name,
            "rows": [
                {"date": "2026-01-01", "value": 0.01},
                {"date": "2026-01-02", "value": 0.02},
            ],
        }"""


def _records_code(workbook_name: str, hash_prefix: str) -> str:
    return f"""from brain_sim.excel import read_excel_expressions
from brain_sim.payloads import build_payload_record

excel_path = DATA_DIR / "{workbook_name}"
alpha_rows = read_excel_expressions(excel_path, sheet_name="alphas")
payload_records = []
for index, alpha in enumerate(alpha_rows, start=1):
    record = build_payload_record(alpha)
    payload_records.append(
        record.__class__(
            row_id=record.row_id,
            alpha_hash=f"{hash_prefix}-{{index}}",
            payload=record.payload,
            metadata=record.metadata,
        )
    )

print(f"Loaded {{len(payload_records)}} payload records from {{excel_path.name}}")
[(record.row_id, record.alpha_hash, record.payload["settings"]["universe"]) for record in payload_records]"""


def build_tutorial_0() -> None:
    _write_notebook(
        TUTORIALS[0],
        [
            markdown_cell("""# Tutorial 0 - Start Here For Beginners

Start here if you are new to WorldQuant BRAIN automation, Python packages, or Jupyter notebooks.

This notebook does not log in, does not call the live BRAIN API, and does not consume simulation quota. It only shows the shape of the workflow: Excel in, run folder out, summary table for review."""),
            markdown_cell("""## 1. The Whole Workflow In One Picture

`brain-sim` has one job:

1. You prepare an Excel file with alpha expressions.
2. The library turns each row into a BRAIN simulation payload.
3. BRAIN returns alpha results.
4. The library saves a review folder with `summary.csv`, raw logs, alpha details, and retry information.

For a beginner, the first goal is not live automation. The first goal is to understand the files."""),
            code_cell("""from __future__ import annotations

import csv
from pathlib import Path

CWD = Path.cwd().resolve()
ROOT = CWD if (CWD / "examples").exists() else CWD.parent if CWD.name == "examples" else CWD
EXAMPLE_DIR = ROOT / "examples"
DATA_DIR = EXAMPLE_DIR / "data"
EXPECTED_DIR = EXAMPLE_DIR / "expected"

print(f"Repo: {ROOT}")
print(f"Sample Excel files live in: {DATA_DIR}")
print(f"Expected output examples live in: {EXPECTED_DIR}")"""),
            markdown_cell("""## 2. Look At A Sample Excel Queue

The minimum useful spreadsheet has one row per alpha expression. The important beginner columns are:

- `id`: your name for the row
- `expression`: the Fast Expression sent to BRAIN
- `region`, `universe`, `delay`: the simulation setting context

Advanced settings can wait until Tutorial 3."""),
            code_cell("""from openpyxl import load_workbook

workbook_path = DATA_DIR / "tutorial_04_live_alphas.xlsx"
workbook = load_workbook(workbook_path, data_only=True)
sheet = workbook["alphas"]
rows = list(sheet.iter_rows(values_only=True))

headers = rows[0]
preview = rows[1:3]
print(headers)
for row in preview:
    print(row[:6])"""),
            markdown_cell("""## 3. Look At The Result Table

After a run, the main file to review is `summary.csv`.

The beginner columns are:

- `row_id`: which Excel row this result came from
- `status`: complete, failed, timed out, or skipped duplicate
- `alpha_id`: the BRAIN alpha identifier when available
- `sharpe`, `fitness`, `returns`, `turnover`, `drawdown`: review metrics
- `error`: what to inspect when a row did not complete"""),
            code_cell("""summary_path = EXPECTED_DIR / "tutorial_04_live_offline_summary.csv"
with summary_path.open(newline="", encoding="utf-8") as f:
    summary_rows = list(csv.DictReader(f))

simple_columns = ["row_id", "status", "alpha_id", "sharpe", "fitness", "turnover", "error"]
for row in summary_rows:
    print({column: row[column] for column in simple_columns})"""),
            markdown_cell("""## 4. The First Real Commands

When you are ready for live BRAIN, do this from a terminal, not from this beginner notebook:

```bash
brain-sim login --print-link --credentials-file ~/.brain_credentials
brain-sim simulate-excel examples/data/tutorial_04_live_alphas.xlsx --sheet alphas --batch-size 4
```

Live simulation consumes BRAIN quota. Learn the offline tutorials first, then run a very small live batch."""),
            markdown_cell("""## 5. Where To Go Next

Recommended path for beginners:

1. Tutorial 1 - Installation And Project Tour
2. Tutorial 3 - Excel Alpha Queue And Payloads
3. Tutorial 4 - Live Excel Batch Simulation
4. Tutorial 6 - Duplicate Cache And Re-Runs
5. Tutorial 7 - Results Raw Logs And Recordsets

Read Tutorial 2 only when you are ready to log in. Read Tutorial 5 after you have seen a timeout or retry queue. Read Tutorial 8 when you want to write custom Python automation."""),
        ],
    )


def build_tutorial_1() -> None:
    _write_notebook(
        TUTORIALS[1],
        [
            markdown_cell("""# Tutorial 1 - Installation And Project Tour

This tutorial introduces the installed package, CLI entry points, repository layout, and safety boundaries before any live BRAIN work."""),
            markdown_cell("""## 1. Install Editable Mode

From a terminal, install the package with development dependencies:

```bash
python -m venv .venv
. .venv/bin/activate
python -m pip install -e ".[dev]"
```

The notebook cells below inspect the current checkout without changing files."""),
            code_cell(COMMON_SETUP),
            markdown_cell("""## 2. Inspect CLI Help And Version

The CLI exposes `login` and `simulate-excel`. Use CLI workflows for standard runs and Python APIs when you need custom automation."""),
            code_cell("""from brain_sim import __version__
from brain_sim.cli import build_parser

parser = build_parser()
print(f"brain-sim version: {__version__}")
print("\\n".join(parser.format_help().splitlines()[:14]))"""),
            markdown_cell("""## 3. Repository Layout

Core package code lives in `src/brain_sim`, notebooks and sample assets live in `examples`, and tests live in `tests`."""),
            code_cell("""for path in ["src/brain_sim", "examples", "examples/data", "examples/expected", "tests"]:
    target = ROOT / path
    print(f"{path:24} exists={target.exists()}")"""),
            markdown_cell("""## 4. Safety Paths

Runtime cookies, run folders, virtual environments, and credentials should not be committed. `examples/runs/` is also ignored because notebooks write local artifacts there."""),
            code_cell("""gitignore = (ROOT / ".gitignore").read_text(encoding="utf-8").splitlines()
for pattern in [".brain_sim/", "runs/", "examples/runs/", ".venv/", "dist/"]:
    print(f"{pattern:16} ignored={pattern in gitignore}")"""),
            markdown_cell("""## 5. CLI Or Python API

Use `brain-sim login` and `brain-sim simulate-excel` for repeatable operational runs. Use the Python API when you need custom queue building, additional filtering, or integration with another research system."""),
        ],
    )


def build_tutorial_2() -> None:
    _write_notebook(
        TUTORIALS[2],
        [
            markdown_cell("""# Tutorial 2 - Login And Persona Verification

This tutorial explains credential loading, Persona verification, cookie storage, and SMTP notification without contacting the live BRAIN API."""),
            markdown_cell("""## 1. Credentials File Format

`brain-sim` accepts either a JSON list `[email, password]` or an object with `email` and `password`. Do not commit real credentials."""),
            code_cell(COMMON_SETUP),
            code_cell("""from brain_sim.auth import BrainAuth, load_credentials
from brain_sim.models import AuthChallenge
from brain_sim.notify import MemoryNotifier

login_run_dir = RUNS_DIR / "tutorial_02_login"
shutil.rmtree(login_run_dir, ignore_errors=True)
login_run_dir.mkdir(parents=True)

demo_credentials = login_run_dir / "demo_credentials.json"
demo_credentials.write_text(json.dumps(["researcher@example.com", "REPLACE_ME"]), encoding="utf-8")
email, password = load_credentials(demo_credentials)
print(email, password)"""),
            markdown_cell("""## 2. Persona Verification Link Lifecycle

When BRAIN returns a Persona challenge, the CLI prints the verification link and saves challenge state in `.brain_sim/latest_login_link.json`. This demo writes the same shape into the tutorial run folder."""),
            code_cell("""challenge = AuthChallenge(
    url="https://inquiry.withpersona.com/verify?inquiry-id=demo-inquiry",
    www_authenticate="persona",
    message="WorldQuant BRAIN requires Persona verification.",
    payload={"inquiry": "demo-inquiry", "verification_url": "https://inquiry.withpersona.com/verify?inquiry-id=demo-inquiry"},
)
challenge_path = login_run_dir / "latest_login_link.json"
challenge_path.write_text(json.dumps(challenge.__dict__, indent=2, sort_keys=True), encoding="utf-8")
print(challenge.url)
print(challenge_path)"""),
            markdown_cell("""## 3. Email Notification Mode

SMTP mode sends the same Persona URL to your email. The live CLI reads `BRAIN_SIM_SMTP_HOST`, `BRAIN_SIM_SMTP_PORT`, `BRAIN_SIM_SMTP_FROM`, `BRAIN_SIM_SMTP_USER`, and `BRAIN_SIM_SMTP_PASSWORD`."""),
            code_cell("""notifier = MemoryNotifier()
notifier.send_login_link("me@example.com", challenge.url)
notifier.sent_messages[0]"""),
            markdown_cell("""## 4. Cookie Storage And Reload

After successful authentication, cookies are stored in `.brain_sim/cookies.json`. This demo writes a safe fake cookie and confirms `BrainAuth.load_saved_cookies()` can load it."""),
            code_cell("""cookie_path = login_run_dir / "cookies.json"
cookie_path.write_text(
    json.dumps({"cookies": [{"name": "t", "value": "demo", "domain": "api.worldquantbrain.com", "path": "/", "secure": True, "expires": None}]}, indent=2),
    encoding="utf-8",
)
auth = BrainAuth(cookie_path=cookie_path)
loaded = auth.load_saved_cookies()
print(f"loaded={loaded}")
print(auth.session.cookies.get("t"))"""),
            markdown_cell("""## 5. Live Login Commands

Run these from a terminal when you are ready to authenticate for real:

```bash
brain-sim login --print-link --credentials-file ~/.brain_credentials
brain-sim login --notify-email "me@example.com" --credentials-file ~/.brain_credentials
```

The second command requires SMTP environment variables. Complete Persona in the browser, then run login again to save cookies."""),
        ],
    )


def build_tutorial_3() -> None:
    _write_notebook(
        TUTORIALS[3],
        [
            markdown_cell("""# Tutorial 3 - Excel Alpha Queue And Payloads

This tutorial covers the Excel schema, settings overrides, metadata columns, payload generation, and duplicate identity hashes."""),
            code_cell(COMMON_SETUP),
            markdown_cell("""## 1. Read A Valid Excel Queue

The required column is `expression`. Optional simulation settings override `SimulationSettings`; other columns become metadata."""),
            code_cell(_records_code("tutorial_03_mixed_settings.xlsx", "tutorial-hash")),
            markdown_cell("""## 2. Inspect Settings And Metadata

`theme` is not a BRAIN setting, so it is preserved as row metadata. Numeric and boolean setting columns are coerced before payload creation."""),
            code_cell("""first_alpha = alpha_rows[0]
print(first_alpha.settings)
print(first_alpha.metadata)
print(payload_records[0].payload)"""),
            markdown_cell("""## 3. Payload Hash Identity

The duplicate key is a SHA-256 hash of the full normalized simulation payload. Changing the expression or any setting changes the hash."""),
            code_cell("""from brain_sim.payloads import build_regular_payload, hash_payload, normalize_payload

payload = build_regular_payload(first_alpha)
print(normalize_payload(payload))
print(hash_payload(payload))"""),
            markdown_cell("""## 4. Invalid Excel Schema

Missing `expression` raises `ExcelInputError` before any simulation is submitted."""),
            code_cell("""from brain_sim.excel import ExcelInputError, read_excel_expressions

invalid_path = DATA_DIR / "tutorial_03_invalid_missing_expression.xlsx"
try:
    read_excel_expressions(invalid_path, sheet_name="alphas")
except ExcelInputError as exc:
    print(type(exc).__name__, str(exc))"""),
        ],
    )


def build_tutorial_4() -> None:
    _write_notebook(
        TUTORIALS[4],
        [
            markdown_cell("""# Tutorial 4 - Live Excel Batch Simulation

This tutorial shows the full Excel simulation workflow. Offline cells use a fake client; live cells submit to BRAIN only when `BRAIN_SIM_RUN_LIVE=1`."""),
            code_cell(COMMON_SETUP),
            code_cell(FAKE_CLIENTS),
            markdown_cell("""## 1. Build Payloads From Excel

The same workbook can be used for offline rehearsal and live `simulate-excel` runs."""),
            code_cell(_records_code("tutorial_04_live_alphas.xlsx", "tutorial-hash")),
            markdown_cell("""## 2. Offline Batch Run

The offline run writes the same artifact shape as a live run: raw logs, alpha detail JSON, summary CSV, retry queue, and SQLite cache."""),
            code_cell("""from brain_sim.batch import BatchRunner

run_dir = RUNS_DIR / "tutorial_04_offline_batch"
shutil.rmtree(run_dir, ignore_errors=True)
runner = BatchRunner(CompleteFakeBrainClient(), run_dir)
result = runner.run(payload_records, batch_size=4, poll_timeout_seconds=5)
print(result)

summary_path = run_dir / "summary.csv"
with summary_path.open(newline="", encoding="utf-8") as f:
    summary_rows = list(csv.DictReader(f))
summary_rows"""),
            markdown_cell("""## 3. Live CLI Commands

These commands consume live BRAIN simulation quota. Run them only after `brain-sim login` has saved cookies."""),
            code_cell("""live_commands = [
    ["brain-sim", "simulate-excel", str(DATA_DIR / "tutorial_04_live_alphas.xlsx"), "--sheet", "alphas", "--batch-size", "auto", "--poll-timeout-seconds", "1800"],
    ["brain-sim", "simulate-excel", str(DATA_DIR / "tutorial_04_live_alphas.xlsx"), "--sheet", "alphas", "--batch-size", "8"],
    ["brain-sim", "simulate-excel", str(DATA_DIR / "tutorial_04_live_alphas.xlsx"), "--sheet", "alphas", "--batch-size", "4"],
    ["brain-sim", "simulate-excel", str(DATA_DIR / "tutorial_04_live_alphas.xlsx"), "--sheet", "alphas", "--batch-size", "1"],
]

if RUN_LIVE:
    subprocess.run(live_commands[0], cwd=ROOT, check=False)
else:
    for command in live_commands:
        print(" ".join(command))"""),
        ],
    )


def build_tutorial_5() -> None:
    _write_notebook(
        TUTORIALS[5],
        [
            markdown_cell("""# Tutorial 5 - Batch Fallback, Timeouts, And Retry Queue

This tutorial demonstrates compatible grouping, 8-to-4-to-1 fallback, pending timeouts, and retry queue review."""),
            code_cell(COMMON_SETUP),
            code_cell(FAKE_CLIENTS),
            code_cell(_records_code("tutorial_05_fallback_alphas.xlsx", "fallback-hash")),
            markdown_cell("""## 1. Batch Compatibility

The runner batches only compatible payloads: type, instrument type, region, universe, delay, and language must match."""),
            code_cell("""from brain_sim.batch import iter_allowed_batch_chunks

print([len(chunk) for chunk in iter_allowed_batch_chunks(payload_records, 8)])"""),
            markdown_cell("""## 2. Rejected Multi-Submit Fallback And Timeout

This fake client rejects an 8-item request, accepts 4-item requests, completes the first 4 rows, and times out the second 4 rows."""),
            code_cell("""class FallbackTimeoutFakeClient(CompleteFakeBrainClient):
    def submit(self, payload: dict[str, Any] | list[dict[str, Any]]) -> TutorialSubmitResult:
        payloads = payload if isinstance(payload, list) else [payload]
        if len(payloads) > 4:
            return TutorialSubmitResult(
                status_code=422,
                location="",
                body={"error": "multi-submit payload too large"},
                headers={},
                rate_limit=TutorialRateLimitState(limit=None, remaining=None, reset_seconds=None),
            )
        self.submit_count += 1
        location = "/simulations/fallback-timeout" if self.submit_count == 2 else f"/simulations/fallback-{self.submit_count}"
        self.locations[location] = [f"tutorial-alpha-{index}" for index in range(1 + (self.submit_count - 1) * 4, 1 + self.submit_count * 4)]
        return TutorialSubmitResult(201, location, {"id": location}, {"Location": location}, TutorialRateLimitState(None, None, None))

    def poll(self, location: str, *, timeout_seconds: float) -> TutorialPollResult:
        if location == "/simulations/fallback-timeout":
            return TutorialPollResult(status="pending_timeout", body=None, events=[{"progress": 0.35}])
        return super().poll(location, timeout_seconds=timeout_seconds)


from brain_sim.batch import BatchRunner

run_dir = RUNS_DIR / "tutorial_05_fallback"
shutil.rmtree(run_dir, ignore_errors=True)
runner = BatchRunner(FallbackTimeoutFakeClient(), run_dir)
result = runner.run(payload_records, batch_size="auto", poll_timeout_seconds=1)
print(result)
print((run_dir / "retry_queue.jsonl").read_text(encoding="utf-8"))"""),
            markdown_cell("""## 3. Retry Review

Do not blindly resubmit ambiguous transport errors. Inspect `retry_queue.jsonl`, preserve `simulation_location`, and decide whether to wait longer, fix expressions, or retry only specific rows."""),
        ],
    )


def build_tutorial_6() -> None:
    _write_notebook(
        TUTORIALS[6],
        [
            markdown_cell("""# Tutorial 6 - Duplicate Cache And Re-Runs

This tutorial shows how `simulation_cache.sqlite` prevents duplicate submissions for the same normalized payload."""),
            code_cell(COMMON_SETUP),
            code_cell(FAKE_CLIENTS),
            code_cell(_records_code("tutorial_06_duplicate_alphas.xlsx", "duplicate-hash")),
            markdown_cell("""## 1. First Run Submits

The first run records successful alpha IDs in SQLite."""),
            code_cell("""from brain_sim.batch import BatchRunner
from brain_sim.cache import SimulationCache

base_run_dir = RUNS_DIR / "tutorial_06_duplicates"
shutil.rmtree(base_run_dir, ignore_errors=True)
cache_path = base_run_dir / "simulation_cache.sqlite"

first_runner = BatchRunner(CompleteFakeBrainClient(alpha_prefix="duplicate-alpha"), base_run_dir / "first", cache_path=cache_path)
first_result = first_runner.run(payload_records, batch_size=1, poll_timeout_seconds=5)
print(first_result)
SimulationCache(cache_path).all_rows()"""),
            markdown_cell("""## 2. Second Run Skips

The second run uses the same cache path and writes `skipped_duplicate` rows instead of submitting again."""),
            code_cell("""second_runner = BatchRunner(CompleteFakeBrainClient(alpha_prefix="should-not-submit"), base_run_dir / "second", cache_path=cache_path)
second_result = second_runner.run(payload_records, batch_size=1, poll_timeout_seconds=5)
print(second_result)
with (base_run_dir / "second" / "summary.csv").open(newline="", encoding="utf-8") as f:
    list(csv.DictReader(f))"""),
            markdown_cell("""## 3. What Changes The Hash

Changing expression text or any simulation setting creates a different payload hash. Metadata-only columns do not enter the BRAIN payload hash."""),
        ],
    )


def build_tutorial_7() -> None:
    _write_notebook(
        TUTORIALS[7],
        [
            markdown_cell("""# Tutorial 7 - Results, Raw Logs, And Recordsets

This tutorial reviews the run folder after a completed simulation and demonstrates recordset capture."""),
            code_cell(COMMON_SETUP),
            code_cell(FAKE_CLIENTS),
            code_cell(_records_code("tutorial_07_recordset_alphas.xlsx", "recordset-hash")),
            markdown_cell("""## 1. Run With Recordsets

`--recordset pnl --recordset sharpe` asks the runner to fetch extra alpha recordsets after successful completion."""),
            code_cell("""from brain_sim.batch import BatchRunner

run_dir = RUNS_DIR / "tutorial_07_recordsets"
shutil.rmtree(run_dir, ignore_errors=True)
runner = BatchRunner(CompleteFakeBrainClient(alpha_prefix="recordset-alpha"), run_dir)
result = runner.run(payload_records, batch_size=1, poll_timeout_seconds=5, recordsets=["pnl", "sharpe"])
print(result)
for path in sorted(run_dir.rglob("*")):
    if path.is_file():
        print(path.relative_to(run_dir))"""),
            markdown_cell("""## 2. Inspect Summary, raw JSONL Logs, Alpha Details, And Recordsets

The summary is the review table. Raw JSONL logs preserve submit and poll context. Alpha detail JSON and recordsets support deeper audit work."""),
            code_cell("""summary = list(csv.DictReader((run_dir / "summary.csv").open(newline="", encoding="utf-8")))
raw_submit_path = next((run_dir / "raw").glob("submit-*.jsonl"))
raw_submit_lines = raw_submit_path.read_text(encoding="utf-8").splitlines()
recordset_files = sorted(path.relative_to(run_dir) for path in (run_dir / "recordsets").rglob("*.json"))
print(summary)
print(f"raw submit events: {len(raw_submit_lines)}")
print(recordset_files)"""),
        ],
    )


def build_tutorial_8() -> None:
    _write_notebook(
        TUTORIALS[8],
        [
            markdown_cell("""# Tutorial 8 - Python API Workflow

This tutorial assembles the public Python pieces into a small automation workflow without using the CLI."""),
            code_cell(COMMON_SETUP),
            code_cell(FAKE_CLIENTS),
            markdown_cell("""## 1. Import The API Pieces

Use `BrainAuth` for cookie management, `read_excel_expressions` for input, `build_payload_record` for payloads, `BrainClient` for live HTTP access, `BatchRunner` for orchestration, and `RunStore` for artifacts."""),
            code_cell("""from brain_sim.auth import BrainAuth
from brain_sim.batch import BatchRunner
from brain_sim.client import BrainClient
from brain_sim.excel import read_excel_expressions
from brain_sim.payloads import build_payload_record
from brain_sim.results import RunStore

cookie_path = RUNS_DIR / "tutorial_08_api" / "cookies.json"
auth = BrainAuth(cookie_path=cookie_path)
live_client = BrainClient()
store = RunStore(RUNS_DIR / "tutorial_08_api" / "empty_store")
print(type(auth).__name__, type(live_client).__name__, store.run_dir.exists())"""),
            markdown_cell("""## 2. Build A Custom Offline Automation Function

The same function can accept `BrainClient()` after authentication. In this notebook it receives a fake client."""),
            code_cell("""def run_excel_queue(excel_path: Path, *, client: Any, run_dir: Path) -> dict[str, Any]:
    alpha_rows = read_excel_expressions(excel_path, sheet_name="alphas")
    records = []
    for index, alpha in enumerate(alpha_rows, start=1):
        record = build_payload_record(alpha)
        records.append(record.__class__(record.row_id, f"api-hash-{index}", record.payload, record.metadata))
    runner = BatchRunner(client, run_dir)
    return runner.run(records, batch_size="auto", poll_timeout_seconds=5, recordsets=["pnl"])


api_run_dir = RUNS_DIR / "tutorial_08_api" / "run"
shutil.rmtree(api_run_dir, ignore_errors=True)
result = run_excel_queue(DATA_DIR / "tutorial_08_api_alphas.xlsx", client=CompleteFakeBrainClient(alpha_prefix="api-alpha"), run_dir=api_run_dir)
print(result)
sorted(path.relative_to(api_run_dir) for path in api_run_dir.rglob("*") if path.is_file())"""),
            markdown_cell("""## 3. Optional Live Swap

After `brain-sim login`, you can load saved cookies into a `requests.Session`, construct `BrainClient(session=session)`, and pass that client into the same function. Keep the `BRAIN_SIM_RUN_LIVE=1` guard for any notebook-based live execution."""),
        ],
    )


def build_notebooks() -> None:
    for obsolete in OBSOLETE_NOTEBOOKS:
        (EXAMPLES / obsolete).unlink(missing_ok=True)
    build_tutorial_0()
    build_tutorial_1()
    build_tutorial_2()
    build_tutorial_3()
    build_tutorial_4()
    build_tutorial_5()
    build_tutorial_6()
    build_tutorial_7()
    build_tutorial_8()


def main() -> None:
    build_workbooks()
    build_expected_outputs()
    build_notebooks()


if __name__ == "__main__":
    main()
