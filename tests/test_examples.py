from __future__ import annotations

import csv
import json
import os
import sys
import types
from pathlib import Path
from urllib.parse import unquote

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
EXAMPLES = ROOT / "examples"
DATA = EXAMPLES / "data"
EXPECTED = EXAMPLES / "expected"
EXAMPLES_README = EXAMPLES / "README.md"
ROOT_README = ROOT / "README.md"

TUTORIALS = [
    "Tutorial 1 - Installation And Project Tour.ipynb",
    "Tutorial 2 - Login And Persona Verification.ipynb",
    "Tutorial 3 - Excel Alpha Queue And Payloads.ipynb",
    "Tutorial 4 - Live Excel Batch Simulation.ipynb",
    "Tutorial 5 - Batch Fallback Timeouts And Retry Queue.ipynb",
    "Tutorial 6 - Duplicate Cache And Re-Runs.ipynb",
    "Tutorial 7 - Results Raw Logs And Recordsets.ipynb",
    "Tutorial 8 - Python API Workflow.ipynb",
]

WORKBOOKS = [
    "tutorial_03_mixed_settings.xlsx",
    "tutorial_03_invalid_missing_expression.xlsx",
    "tutorial_04_live_alphas.xlsx",
    "tutorial_05_fallback_alphas.xlsx",
    "tutorial_06_duplicate_alphas.xlsx",
    "tutorial_07_recordset_alphas.xlsx",
    "tutorial_08_api_alphas.xlsx",
]

EXPECTED_FILES = [
    "tutorial_04_live_offline_summary.csv",
    "tutorial_05_fallback_summary.csv",
    "tutorial_06_second_run_summary.csv",
    "tutorial_07_recordset_summary.csv",
]


def _load_notebook(name: str) -> dict:
    return json.loads((EXAMPLES / name).read_text(encoding="utf-8"))


def _notebook_source(name: str) -> str:
    notebook = _load_notebook(name)
    return "\n".join("".join(cell.get("source", [])) for cell in notebook["cells"])


def _code_cell_source(name: str) -> str:
    notebook = _load_notebook(name)
    return "\n".join(
        "".join(cell.get("source", []))
        for cell in notebook["cells"]
        if cell.get("cell_type") == "code"
    )


def test_examples_index_links_all_tutorials_and_assets() -> None:
    text = EXAMPLES_README.read_text(encoding="utf-8")

    for tutorial in TUTORIALS:
        assert tutorial.replace(".ipynb", "") in text
        assert f"]({tutorial.replace(' ', '%20')})" in text
        assert (EXAMPLES / tutorial).exists()

    for workbook in WORKBOOKS:
        assert workbook in text
        assert (DATA / workbook).exists()

    for expected_file in EXPECTED_FILES:
        assert expected_file in text
        assert (EXPECTED / expected_file).exists()

    assert "BRAIN_SIM_RUN_LIVE=1" in text
    assert "offline-safe" in text


def test_root_readme_points_to_tutorial_suite() -> None:
    text = ROOT_README.read_text(encoding="utf-8")

    assert "## Examples" in text
    assert "examples/README.md" in text
    assert "Tutorial 8 - Python API Workflow" in text
    for tutorial in TUTORIALS:
        encoded_target = f"examples/{tutorial.replace(' ', '%20')}"
        assert f"]({encoded_target})" in text
        assert (ROOT / unquote(encoded_target)).exists()
    assert "BRAIN_SIM_RUN_LIVE=1" in text


def test_notebooks_have_required_sections_and_feature_terms() -> None:
    required = {
        TUTORIALS[0]: ["# Tutorial 1 - Installation And Project Tour", "CLI Help", "Safety Paths"],
        TUTORIALS[1]: ["# Tutorial 2 - Login And Persona Verification", "Persona Verification Link", "Email Notification", "Cookie Storage"],
        TUTORIALS[2]: ["# Tutorial 3 - Excel Alpha Queue And Payloads", "Payload Hash", "Invalid Excel Schema"],
        TUTORIALS[3]: ["# Tutorial 4 - Live Excel Batch Simulation", "Live CLI Commands", "batch-size"],
        TUTORIALS[4]: ["# Tutorial 5 - Batch Fallback, Timeouts, And Retry Queue", "8-to-4-to-1", "retry_queue.jsonl"],
        TUTORIALS[5]: ["# Tutorial 6 - Duplicate Cache And Re-Runs", "simulation_cache.sqlite", "skipped_duplicate"],
        TUTORIALS[6]: ["# Tutorial 7 - Results, Raw Logs, And Recordsets", "recordsets", "raw JSONL"],
        TUTORIALS[7]: ["# Tutorial 8 - Python API Workflow", "BrainAuth", "BrainClient", "BatchRunner", "RunStore"],
    }

    for tutorial, terms in required.items():
        source = _notebook_source(tutorial)
        for term in terms:
            assert term in source


def test_live_notebook_cells_are_gated_and_do_not_hardcode_credentials() -> None:
    forbidden_code_fragments = [
        "api_key",
        "secret",
        "smtp-password",
        ".brain_credentials",
    ]

    for tutorial in TUTORIALS:
        code = _code_cell_source(tutorial)
        assert 'RUN_LIVE = os.getenv("BRAIN_SIM_RUN_LIVE") == "1"' in code
        for fragment in forbidden_code_fragments:
            assert fragment not in code

    live_code = _code_cell_source(TUTORIALS[3])
    assert "subprocess.run" in live_code
    assert "if RUN_LIVE:" in live_code


def test_sample_workbooks_have_expected_schema() -> None:
    required_headers = {
        "id",
        "expression",
        "region",
        "universe",
        "delay",
        "decay",
        "neutralization",
        "truncation",
        "nanHandling",
        "language",
        "visualization",
        "theme",
    }

    for workbook_name in WORKBOOKS:
        workbook = load_workbook(DATA / workbook_name, data_only=True)
        sheet = workbook["alphas"]
        rows = list(sheet.iter_rows(values_only=True))
        headers = set(rows[0])
        if "invalid_missing_expression" in workbook_name:
            assert "expression" not in headers
            continue
        assert required_headers.issubset(headers)
        assert len(rows) >= 3


def test_expected_csv_outputs_have_known_statuses() -> None:
    with (EXPECTED / "tutorial_05_fallback_summary.csv").open(newline="", encoding="utf-8") as f:
        fallback_rows = list(csv.DictReader(f))
    assert [row["status"] for row in fallback_rows[:4]] == ["complete"] * 4
    assert [row["status"] for row in fallback_rows[4:]] == ["pending_timeout"] * 4
    assert all(row["simulation_location"] == "/simulations/fallback-timeout" for row in fallback_rows[4:])

    with (EXPECTED / "tutorial_06_second_run_summary.csv").open(newline="", encoding="utf-8") as f:
        duplicate_rows = list(csv.DictReader(f))
    assert [row["status"] for row in duplicate_rows] == ["skipped_duplicate", "skipped_duplicate"]


def test_gitignore_protects_runtime_artifacts() -> None:
    text = (ROOT / ".gitignore").read_text(encoding="utf-8")
    for pattern in [".brain_sim/", "runs/", "examples/runs/", ".venv/", "dist/"]:
        assert pattern in text


def test_all_tutorial_code_cells_execute_offline(monkeypatch) -> None:
    monkeypatch.delenv("BRAIN_SIM_RUN_LIVE", raising=False)
    monkeypatch.chdir(ROOT)

    for tutorial in TUTORIALS:
        notebook = _load_notebook(tutorial)
        module = types.ModuleType(f"__{Path(tutorial).stem.replace(' ', '_')}__")
        namespace: dict[str, object] = module.__dict__
        monkeypatch.setitem(sys.modules, module.__name__, module)

        for index, cell in enumerate(notebook["cells"]):
            if cell.get("cell_type") != "code":
                continue
            source = "".join(cell.get("source", []))
            compiled = compile(source, f"{tutorial}:cell-{index}", "exec")
            exec(compiled, namespace)

    assert (EXAMPLES / "runs" / "tutorial_04_offline_batch" / "summary.csv").exists()
    assert (EXAMPLES / "runs" / "tutorial_05_fallback" / "retry_queue.jsonl").exists()
    assert (EXAMPLES / "runs" / "tutorial_06_duplicates" / "second" / "summary.csv").exists()
    assert (EXAMPLES / "runs" / "tutorial_07_recordsets" / "recordsets").exists()


def test_offline_notebook_outputs_match_expected_csvs() -> None:
    comparisons = [
        (
            EXAMPLES / "runs" / "tutorial_04_offline_batch" / "summary.csv",
            EXPECTED / "tutorial_04_live_offline_summary.csv",
            ["row_id", "status", "alpha_id", "simulation_location", "sharpe", "fitness", "returns", "turnover", "drawdown"],
        ),
        (
            EXAMPLES / "runs" / "tutorial_05_fallback" / "summary.csv",
            EXPECTED / "tutorial_05_fallback_summary.csv",
            ["row_id", "status", "alpha_id", "simulation_location", "error"],
        ),
        (
            EXAMPLES / "runs" / "tutorial_06_duplicates" / "second" / "summary.csv",
            EXPECTED / "tutorial_06_second_run_summary.csv",
            ["row_id", "status", "alpha_id"],
        ),
        (
            EXAMPLES / "runs" / "tutorial_07_recordsets" / "summary.csv",
            EXPECTED / "tutorial_07_recordset_summary.csv",
            ["row_id", "status", "alpha_id", "simulation_location", "sharpe", "fitness"],
        ),
    ]

    for actual_path, expected_path, fields in comparisons:
        with actual_path.open(newline="", encoding="utf-8") as f:
            actual_rows = list(csv.DictReader(f))
        with expected_path.open(newline="", encoding="utf-8") as f:
            expected_rows = list(csv.DictReader(f))

        assert [
            {field: row[field] for field in fields}
            for row in actual_rows
        ] == [
            {field: row[field] for field in fields}
            for row in expected_rows
        ]
